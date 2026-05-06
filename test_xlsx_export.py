from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from models.leg import DuploIndexadorLeg, PreLeg
from models.swap import Swap
from utils.calendars import B3Calendar
from utils.xlsx_export import (
    NUMBER_FORMAT,
    PERCENT_FORMAT,
    USD_FORMAT,
    build_calculation_memory_xlsx,
)


def test_build_calculation_memory_xlsx_contains_expected_sheets_and_summary():
    calendar = B3Calendar()
    start = date(2024, 1, 2)
    end = date(2024, 7, 2)
    active_leg = PreLeg(1_000_000, start, end, 0.12, cotacao_cliente=5.0)
    passive_leg = PreLeg(1_000_000, start, end, 0.10, cotacao_cliente=5.0)
    results = Swap(active_leg, passive_leg).calculate_net_value(calendar)

    xlsx_bytes = build_calculation_memory_xlsx(
        notional=1_000_000,
        start_date=start,
        end_date=end,
        du=calendar.business_days(start, end),
        dc=calendar.calendar_days(start, end),
        type_active="Pré-Fixada",
        params_active={"type": "Pré-Fixada", "rate": 0.12, "cotacao": 5.0},
        leg_active=active_leg,
        type_passive="Pré-Fixada",
        params_passive={"type": "Pré-Fixada", "rate": 0.10, "cotacao": 5.0},
        leg_passive=passive_leg,
        results=results,
    )

    workbook = load_workbook(BytesIO(xlsx_bytes), data_only=False)

    assert workbook.sheetnames == [
        "Resumo",
        "Entradas",
        "Ponta Ativa",
        "Ponta Passiva",
        "Fórmulas",
        "Dados",
    ]
    assert workbook["Resumo"]["A1"].value == "Memória de Cálculo do Swap"
    assert workbook["Resumo"]["A10"].value == "Ponta"
    assert workbook["Resumo"]["B11"].value == "Pré-Fixada"
    assert workbook["Fórmulas"]["A1"].value == "Biblioteca de Fórmulas"


def test_build_calculation_memory_xlsx_formats_usd_rates_and_quotes_by_row_label():
    calendar = B3Calendar()
    start = date(2024, 1, 2)
    end = date(2024, 7, 2)
    active_leg = PreLeg(1_000_000, start, end, 0.12, cotacao_cliente=5.0)
    passive_leg = PreLeg(1_000_000, start, end, 0.10, cotacao_cliente=5.0)
    results = Swap(active_leg, passive_leg).calculate_net_value(calendar)

    xlsx_bytes = build_calculation_memory_xlsx(
        notional=1_000_000,
        start_date=start,
        end_date=end,
        du=calendar.business_days(start, end),
        dc=calendar.calendar_days(start, end),
        type_active="Pré-Fixada",
        params_active={"type": "Pré-Fixada", "rate": 0.12, "cotacao": 5.0},
        leg_active=active_leg,
        type_passive="Pré-Fixada",
        params_passive={"type": "Pré-Fixada", "rate": 0.10, "cotacao": 5.0},
        leg_passive=passive_leg,
        results=results,
    )

    workbook = load_workbook(BytesIO(xlsx_bytes), data_only=False)

    assert workbook["Resumo"]["B6"].number_format == USD_FORMAT
    assert workbook["Ponta Ativa"]["B5"].number_format == USD_FORMAT
    assert workbook["Ponta Passiva"]["B5"].number_format == USD_FORMAT
    assert workbook["Entradas"]["C6"].number_format == PERCENT_FORMAT
    assert workbook["Entradas"]["C7"].number_format == NUMBER_FORMAT


def test_build_calculation_memory_xlsx_includes_duplo_indexador_components_and_market_data():
    calendar = B3Calendar()
    start = date(2024, 1, 2)
    end = date(2025, 1, 2)
    active_leg = DuploIndexadorLeg(
        1_000_000,
        start,
        end,
        cotacao_cliente=5.0,
        cotacao_atual=6.5,
        spread_pre=0.05,
        spread_vc=0.10,
        cap=5.5,
        cap_target="vc",
        use_vc_contra=False,
    )
    passive_leg = PreLeg(1_000_000, start, end, 0.10, cotacao_cliente=5.0)
    results = Swap(active_leg, passive_leg).calculate_net_value(calendar)

    xlsx_bytes = build_calculation_memory_xlsx(
        notional=1_000_000,
        start_date=start,
        end_date=end,
        du=calendar.business_days(start, end),
        dc=calendar.calendar_days(start, end),
        type_active="Duplo Indexador",
        params_active={
            "type": "Duplo Indexador",
            "cotacao_cliente": 5.0,
            "cotacao_atual": 6.5,
            "spread_pre": 0.05,
            "spread_vc": 0.10,
            "cap": 5.5,
            "cap_target": "vc",
        },
        leg_active=active_leg,
        type_passive="Pré-Fixada",
        params_passive={"type": "Pré-Fixada", "rate": 0.10, "cotacao": 5.0},
        leg_passive=passive_leg,
        results=results,
    )

    workbook = load_workbook(BytesIO(xlsx_bytes), data_only=False)
    active_sheet_values = [cell.value for row in workbook["Ponta Ativa"].iter_rows() for cell in row]
    formulas_values = [cell.value for row in workbook["Fórmulas"].iter_rows() for cell in row]

    assert "Componente VC antes do CAP" in active_sheet_values
    assert "Componente VC após CAP" in active_sheet_values
    assert "CAP aplicado" in active_sheet_values
    assert "max(Componente Pré, Componente VC após CAP)" in formulas_values
    assert workbook["Dados"]["A1"].value == "Dados Resolvidos e Fontes"

    active_sheet = workbook["Ponta Ativa"]
    cap_row = next(row for row in active_sheet.iter_rows() if row[4].value == "CAP aplicado")
    assert cap_row[5].number_format == NUMBER_FORMAT
