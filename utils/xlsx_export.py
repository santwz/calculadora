from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BRL_FORMAT = 'R$ #,##0.00;[Red]-R$ #,##0.00'
USD_FORMAT = '$ #,##0.00'
DATE_FORMAT = 'yyyy-mm-dd'
PERCENT_FORMAT = '0.0000%'
NUMBER_FORMAT = '#,##0.000000'


FORMULA_LIBRARY = [
    ["Tipo", "Fórmula", "Descrição"],
    ["Pré Linear", "P * c_cli * s * dias/base + A*c_cli", "Juros simples sobre principal convertido para BRL."],
    ["Pré Exponencial", "P * c_cli * ((1+s)^(dias/base)-1) + A*c_cli", "Juros compostos sobre principal convertido para BRL."],
    ["CDI + Spread", "P * c_cli * (F_DI * (1+s)^(DU/252)-1) + A*c_cli", "Fator DI acumulado com spread anual."],
    ["CDI Percentual", "P * c_cli * (F_%CDI-1) + A*c_cli", "Fator CDI ponderado diariamente pelo percentual."],
    ["VC Parte", "P * c_final * s * DC/360 + A*(c_final/c_inicial)", "Juros cambiais; FV inclui principal em c_final."],
    ["VC Contra", "P * c_cap * s * DC/360 + A*(c_cap/c_inicial)", "Usa c_cap = min(c_final, CAP) quando aplicável."],
    ["IPCA Capitalizado", "P*c_cli*F_IPCA*(1+s)^(DU/252) - P*c_cli + A*c_cli*F_IPCA", "Principal corrigido pelo IPCA e juro real."],
    ["IPCA Não Capitalizado", "P*c_cli*(F_IPCA*(1+s)^(DU/252)-1) + A*c_cli", "IPCA e juro real aplicados ao fator total."],
    ["SOFR", "P*c_final*(spread+r_SOFR)*DC/360 + A*(c_final/c_inicial)", "FV inclui principal em c_final; r_SOFR vem do SOFR Index."],
    ["Duplo Indexador", "max(Componente Pré, Componente VC após CAP)", "Escolhe o maior entre Pré e VC após regra de CAP."],
]


def build_calculation_memory_xlsx(
    *,
    currency,
    notional,
    amortizacao,
    start_date,
    end_date,
    du,
    dc,
    type_active,
    params_active,
    leg_active,
    type_passive,
    params_passive,
    leg_passive,
    results,
) -> bytes:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    sheets = {
        "Resumo": wb.create_sheet("Resumo"),
        "Entradas": wb.create_sheet("Entradas"),
        "Ponta Ativa": wb.create_sheet("Ponta Ativa"),
        "Ponta Passiva": wb.create_sheet("Ponta Passiva"),
        "Fórmulas": wb.create_sheet("Fórmulas"),
        "Dados": wb.create_sheet("Dados"),
    }

    _build_summary(
        sheets["Resumo"],
        currency,
        notional,
        amortizacao,
        start_date,
        end_date,
        du,
        dc,
        type_active,
        type_passive,
        results,
    )
    _build_inputs(sheets["Entradas"], params_active, params_passive)
    _build_leg_sheet(sheets["Ponta Ativa"], "Ponta Ativa", type_active, params_active, leg_active, results, "long")
    _build_leg_sheet(sheets["Ponta Passiva"], "Ponta Passiva", type_passive, params_passive, leg_passive, results, "short")
    _build_formulas(sheets["Fórmulas"])
    _build_market_data(sheets["Dados"], type_active, params_active, leg_active, type_passive, params_passive, leg_passive)

    for ws in wb.worksheets:
        _finish_sheet(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _build_summary(ws, currency, notional, amortizacao, start_date, end_date, du, dc, type_active, type_passive, results):
    _title(ws, "Memória de Cálculo do Swap", "Resumo executivo do cálculo auditável")
    rows = [
        ["Gerado em", datetime.now()],
        ["Data Início", start_date],
        ["Data Vencimento", end_date],
        ["Moeda", currency],
        [f"Notional {currency}", notional],
        [f"Amortização {currency}", amortizacao],
        ["Dias Úteis (DU)", du],
        ["Dias Corridos (DC)", dc],
    ]
    _write_key_values(ws, rows, start_row=3)

    table = [
        ["Ponta", "Tipo", "Valor Futuro (BRL)", "Fluxo (BRL)"],
        ["Ativa", type_active, results["fv_long"], results["flow_long"]],
        ["Passiva", type_passive, results["fv_short"], results["flow_short"]],
        ["Ajuste Líquido", "Ativa - Passiva", results["net_value"], None],
    ]
    _write_table(ws, table, start_row=12, start_col=1)


def _build_inputs(ws, params_active, params_passive):
    _title(ws, "Entradas Informadas", "Parâmetros capturados na interface antes da criação das pontas")
    rows = [["Ponta", "Campo", "Valor"]]
    rows.extend(_params_rows("Ativa", params_active))
    rows.extend(_params_rows("Passiva", params_passive))
    _write_table(ws, rows, start_row=4, start_col=1)


def _build_leg_sheet(ws, title, leg_type, params, leg, results, side):
    _title(ws, title, f"Memória detalhada da {title.lower()}")
    fv_key = "fv_long" if side == "long" else "fv_short"
    flow_key = "flow_long" if side == "long" else "flow_short"

    resolved = _resolved_leg_rows(leg_type, leg)
    _write_table(ws, [["Dado", "Valor", "Fonte"]] + resolved, start_row=4, start_col=1)

    formula_text = _formula_for_type(leg_type)
    calc_rows = [
        ["Item", "Valor"],
        ["Tipo", leg_type],
        ["Fórmula usada", formula_text],
        ["Valor futuro (BRL)", results[fv_key]],
        ["Fluxo líquido (BRL)", results[flow_key]],
    ]
    calc_rows.extend(_duplo_rows(leg) if leg_type == "Duplo Indexador" else [])
    _write_table(ws, calc_rows, start_row=4, start_col=5)


def _build_formulas(ws):
    _title(ws, "Biblioteca de Fórmulas", "Referência textual das fórmulas utilizadas pelo motor de cálculo")
    _write_table(ws, FORMULA_LIBRARY, start_row=4, start_col=1)


def _build_market_data(ws, type_active, params_active, leg_active, type_passive, params_passive, leg_passive):
    _title(ws, "Dados Resolvidos e Fontes", "Dados manuais e automáticos efetivamente usados no cálculo")
    rows = [["Ponta", "Dado", "Valor", "Fonte"]]
    rows.extend(_market_rows("Ativa", type_active, params_active, leg_active))
    rows.extend(_market_rows("Passiva", type_passive, params_passive, leg_passive))
    _write_table(ws, rows, start_row=4, start_col=1)


def _params_rows(side, params):
    rows = []
    for key, value in params.items():
        rows.append([side, _pretty_key(key), _safe_value(value)])
    return rows


def _resolved_leg_rows(leg_type, leg):
    rows = [
        ["Moeda", getattr(leg, "currency", None), "Modelo"],
        [f"Notional {getattr(leg, 'currency', 'USD')}", getattr(leg, "notional", None), "Modelo"],
        ["Data Início", getattr(leg, "start_date", None), "Modelo"],
        ["Data Vencimento", getattr(leg, "end_date", None), "Modelo"],
    ]

    attr_map = {
        "Pré-Fixada": ["rate", "method", "base", "cotacao_cliente", "amortizacao"],
        "CDI": ["cdi_factor", "spread", "percent", "cotacao_cliente", "amortizacao"],
        "CDI Percentual": ["cdi_factor", "percent", "cotacao_cliente", "amortizacao"],
        "Dólar (VC)": ["spot_start", "spot_end", "coupon", "cap", "use_contra", "amortizacao_usd"],
        "Moeda (VC)": ["spot_start", "spot_end", "coupon", "cap", "use_contra", "amortizacao_usd"],
        "IPCA": ["vna_start", "vna_end", "coupon", "capitalizado", "cotacao_cliente", "amortizacao"],
        "SOFR": ["sofr_index_start", "sofr_index_end", "spot_start", "spot_end", "coupon", "amortizacao_usd"],
        "Duplo Indexador": [
            "cotacao_cliente",
            "cotacao_atual",
            "spread_pre",
            "spread_vc",
            "cap",
            "cap_target",
            "use_vc_contra",
            "amortizacao_usd",
        ],
    }

    for attr in attr_map.get(leg_type, []):
        rows.append([_pretty_key(attr), getattr(leg, attr, None), "Modelo"])

    return rows


def _duplo_rows(leg):
    result = getattr(leg, "resultado", None)
    if not result:
        return []

    return [
        ["Componente Pré", result.get("componente_pre")],
        ["Componente VC antes do CAP", result.get("componente_vc_antes_cap")],
        ["Componente VC após CAP", result.get("componente_vc_apos_cap")],
        ["CAP aplicado", result.get("cap_aplicado")],
        ["Componente escolhido", result.get("componente_escolhido")],
    ]


def _market_rows(side, leg_type, params, leg):
    rows = []
    source = "Manual"

    if leg_type in ["CDI", "CDI Percentual"]:
        source = "BCB SGS 12" if params.get("auto_cdi") else "Manual"
        rows.append([side, "Fator CDI", getattr(leg, "cdi_factor", None), source])
    elif leg_type in ["Dólar (VC)", "Moeda (VC)"]:
        rows.append([side, "Cotação Inicial", getattr(leg, "spot_start", None), "Manual"])
        auto_source = "BCB PTAX" if params.get("auto_ptax") else "Manual"
        rows.append([side, "Cotação Final", getattr(leg, "spot_end", None), auto_source])
    elif leg_type == "IPCA":
        if params.get("auto_ipca") and params.get("manual_ipca_start"):
            start_source = "Manual"
            end_source = "IBGE SIDRA/BCB"
        else:
            start_source = "IBGE SIDRA/BCB" if params.get("auto_ipca") else "Manual"
            end_source = start_source
        rows.append([side, "VNA Inicial", getattr(leg, "vna_start", None), start_source])
        rows.append([side, "VNA Final", getattr(leg, "vna_end", None), end_source])
    elif leg_type == "SOFR":
        rows.append([side, "SOFR Index Inicial", getattr(leg, "sofr_index_start", None), "New York Fed SOFRAI"])
        rows.append([side, "SOFR Index Final", getattr(leg, "sofr_index_end", None), "New York Fed SOFRAI"])
        rows.append([side, "Cotação Inicial", getattr(leg, "spot_start", None), "Manual"])
        rows.append([side, "Cotação Final", getattr(leg, "spot_end", None), "Manual"])
    elif leg_type == "Duplo Indexador":
        rows.append([side, "Cotação Cliente", getattr(leg, "cotacao_cliente", None), "Manual"])
        rows.append([side, "Cotação Atual", getattr(leg, "cotacao_atual", None), "Manual"])
        rows.append([side, "CAP", getattr(leg, "cap", None), "Manual" if getattr(leg, "cap_target", "none") == "vc" else "Não aplicado"])
    else:
        rows.append([side, "Dados de mercado", "N/A", source])

    return rows


def _formula_for_type(leg_type):
    formulas = {
        "Pré-Fixada": "P * c_cli * ((1+s)^(dias/base)-1) + A*c_cli",
        "CDI": "P * c_cli * (F_DI * (1+s)^(DU/252)-1) + A*c_cli",
        "CDI Percentual": "P * c_cli * (F_%CDI-1) + A*c_cli",
        "Dólar (VC)": "FV = P*c_final + P*c_final*s*DC/360 + A*(c_final/c_inicial)",
        "Moeda (VC)": "FV = P*c_final + P*c_final*s*DC/360 + A*(c_final/c_inicial)",
        "IPCA": "P*c_cli*F_IPCA*(1+s)^(DU/252) - P*c_cli + A*c_cli*F_IPCA",
        "SOFR": "FV = P*c_final + P*c_final*(spread+r_SOFR)*DC/360 + A*(c_final/c_inicial)",
        "Duplo Indexador": "max(Componente Pré, Componente VC após CAP)",
    }
    return formulas.get(leg_type, "Ver biblioteca de fórmulas")


def _title(ws, title, subtitle):
    ws["A1"] = title
    ws["A2"] = subtitle
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")
    ws["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, color="5B6770")


def _write_key_values(ws, rows, start_row):
    for index, row in enumerate(rows, start=start_row):
        ws.cell(index, 1, row[0])
        ws.cell(index, 2, _safe_value(row[1]))
        ws.cell(index, 1).font = Font(bold=True, color="1F4E78")
        _format_value_cell(ws.cell(index, 2), row[0])


def _write_table(ws, rows, start_row, start_col):
    thin = Side(style="thin", color="D9E2F3")
    headers = rows[0] if rows else []
    for row_index, row in enumerate(rows, start=start_row):
        for col_offset, value in enumerate(row):
            col_index = start_col + col_offset
            cell = ws.cell(row_index, col_index, _safe_value(value))
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index == start_row:
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
                cell.font = Font(bold=True, color="1F4E78")
            _format_value_cell(cell, _table_cell_format_label(headers, row, col_offset))


def _table_cell_format_label(headers, row, col_offset):
    header = headers[col_offset] if col_offset < len(headers) else ""
    if str(header).strip().lower() != "valor":
        return header

    for semantic_header in ["Campo", "Dado", "Item"]:
        if semantic_header in headers:
            semantic_index = headers.index(semantic_header)
            if semantic_index < len(row):
                return row[semantic_index]

    return row[0] if row else header


def _style_currency_range(ws, start_row, start_col, row_count, cols):
    for row in range(start_row, start_row + row_count):
        for col in cols:
            cell = ws.cell(row, col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = BRL_FORMAT


def _format_value_cell(cell, label):
    if isinstance(cell.value, datetime):
        cell.number_format = "yyyy-mm-dd hh:mm"
    elif isinstance(cell.value, date):
        cell.number_format = DATE_FORMAT
    elif isinstance(cell.value, (int, float)):
        label_text = str(label).lower()
        if "usd" in label_text:
            cell.number_format = USD_FORMAT
        elif any(word in label_text for word in ["valor", "fluxo", "ajuste", "componente"]):
            cell.number_format = BRL_FORMAT
        elif any(word in label_text for word in ["rate", "spread", "coupon", "percent", "taxa", "cupom", "juro"]):
            cell.number_format = PERCENT_FORMAT
        else:
            cell.number_format = NUMBER_FORMAT


def _finish_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    widths = {
        "A": 24,
        "B": 24,
        "C": 24,
        "D": 18,
        "E": 28,
        "F": 32,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 22


def _safe_value(value):
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool, date, datetime)):
        return value
    return str(value)


def _pretty_key(key):
    labels = {
        "type": "Tipo",
        "currency": "Moeda",
        "moeda": "Moeda",
        "rate": "Taxa",
        "method": "Método",
        "base": "Base",
        "cotacao": "Cotação",
        "cotacao_cliente": "Cotação Cliente",
        "cotacao_atual": "Cotação Atual",
        "spread": "Spread",
        "spread_pre": "Spread Pré",
        "spread_vc": "Spread VC",
        "spot_start": "Cotação Inicial",
        "spot_end": "Cotação Final",
        "coupon": "Cupom",
        "cap": "CAP",
        "cap_target": "Aplicação do CAP",
        "use_contra": "Usar Contra-Parte",
        "use_vc_contra": "VC Contra-Parte",
        "cdi_factor": "Fator CDI",
        "percent": "% do CDI",
        "auto_cdi": "CDI Automático",
        "auto_ptax": "PTAX Final Automática",
        "auto_ipca": "IPCA Automático",
        "manual_ipca_start": "Cadastrar IPCA Inicial",
        "ipca_months_back": "Meses para trás IPCA",
        "vna_start": "VNA Inicial",
        "vna_end": "VNA Final",
        "capitalizado": "Capitalizado",
        "amortizacao": "Amortização",
        "amortizacao_usd": "Amortização",
    }
    return labels.get(str(key), str(key).replace("_", " ").title())
